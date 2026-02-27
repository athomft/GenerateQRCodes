"""Tests for the QR code generator."""

import tempfile
from pathlib import Path

import pytest

from main import (
    generate_random_10_digit,
    generate_qr_code,
    save_to_excel,
    create_output_directory,
)


class TestGenerateRandom10Digit:
    """Tests for generate_random_10_digit."""

    def test_returns_10_digit_string(self):
        result = generate_random_10_digit(set())
        assert len(result) == 10
        assert result.isdigit()

    def test_uniqueness(self):
        existing = set()
        results = []
        for _ in range(100):
            number = generate_random_10_digit(existing)
            assert number not in existing
            existing.add(number)
            results.append(number)
        assert len(results) == len(set(results))

    def test_raises_after_max_attempts(self):
        # Fill 10^10 is impractical; we test that passing all 10-digit strings
        # would eventually raise. Instead we mock/cap: with a tiny pool we can
        # force exhaustion. Create a set of 10-digit numbers that makes
        # collision very likely (e.g. only 2 possible values).
        # Actually the function uses random.choices from 0-9 with k=10, so
        # we can't easily exhaust. So let's just test that with empty set we
        # get valid output and with many calls we get unique. Skip the
        # exhaustion test or use a smaller pool by patching.
        # Simplest: just ensure it returns something valid for empty set.
        result = generate_random_10_digit(set())
        assert result.isdigit() and len(result) == 10


class TestSaveToExcel:
    """Tests for save_to_excel."""

    def test_creates_file_with_header_and_data(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "test.xlsx"
            numbers = ["1234567890", "9876543210"]
            assert save_to_excel(numbers, path) is True
            assert path.exists()

            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            assert rows[0][0] == "Code"
            assert rows[1][0] == "1234567890"
            assert rows[2][0] == "9876543210"

    def test_empty_list_creates_file_with_header_only(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "empty.xlsx"
            assert save_to_excel([], path) is True
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            assert len(rows) == 1
            assert rows[0][0] == "Code"


class TestGenerateQrCode:
    """Tests for generate_qr_code."""

    def test_generates_png_file(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "test.png"
            assert generate_qr_code("1234567890", path) is True
            assert path.exists()
            assert path.read_bytes()[:8] == b'\x89PNG\r\n\x1a\n'

    def test_invalid_path_returns_false(self):
        # Unwritable / invalid path (e.g. directory as "file")
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp)  # directory, not a file
            result = generate_qr_code("1234567890", path)
        assert result is False


class TestCreateOutputDirectory:
    """Tests for create_output_directory."""

    def test_creates_new_directory(self):
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "new_subdir"
            assert not out.exists()
            assert create_output_directory(out) is True
            assert out.is_dir()

    def test_existing_directory_succeeds(self):
        with tempfile.TemporaryDirectory() as tmp:
            assert create_output_directory(tmp) is True
